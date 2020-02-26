import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import shutil
import logging
import datetime
import time
import csv
import send_email
import io


# list of directories to process - maintained by NPP team
######## Test ########
# output_directory = os.path.join("L:\\", "SAP_Return_Tickets")

######## Production #########
output_directory = os.path.join("O:\\", "Non-Periodical", "Accounts Receivable", "Templates",
                                "SAP Uploads", "Mr. Roboto")

log_file = os.path.join(output_directory, "Return_Ticket_log.log")

logging.basicConfig(filename=log_file, level=logging.DEBUG, format="%(asctime)s - %(levelname)s: %(message)s",
                    datefmt="%m/%d/%Y %I:%M:%S %p")

logging.info("SAP return ticket process has started")

# use the current date to generate the filename for the output file
current_date = datetime.datetime.now()
date = current_date.strftime("%m/%d/%Y")
year = current_date.strftime("%Y")
big_year = current_date.strftime("%y")
month = current_date.strftime("%m")
day = current_date.strftime("%d")

print("Month: {} Day: {} Year: {}".format(month, day, year))

# gather statement period
days_to_statement = 7 - int(datetime.datetime.now().strftime('%w'))
stmt_date = datetime.datetime.now() + datetime.timedelta(days_to_statement)
stmt_year = stmt_date.strftime('%Y')
stmt_month = stmt_date.strftime('%m')
stmt_day = stmt_date.strftime('%d')

print("Statement Date: {}/{}/{}".format(stmt_month, stmt_day, stmt_year))
logging.info("Statement Date: {}/{}/{}".format(stmt_month, stmt_day, stmt_year))

# create directory for output files
output_file_directory = os.path.join(output_directory, "{}.{}.{}".format(month, day, year))
if not os.path.isdir(output_file_directory):
    logging.info("Output directory created: {}".format(output_file_directory))
    os.mkdir(output_file_directory)
else:
    logging.info("Output directory already available: {}".format(output_file_directory))

# read the product lookup tables to memory
product_lookup = []
lookup = load_workbook(os.path.join(output_directory, 'Mr Roboto Lookup.xlsx'))
lookup_wb = lookup.get_active_sheet()
for row in range(2, lookup_wb.max_row + 1):
    upc = lookup_wb.cell(row=row, column=1).value
    item = lookup_wb.cell(row=row, column=2).value
    product_lookup.append([upc, item])

logging.info("Product lookup table loaded")
# empty list to store files that have been processes
files_processed = []
errors = []

# create empty workbook for process output
wb_out = Workbook()
ws_publix = wb_out.active
ws_publix.title = "Publix"
publix_output_row = 1
# add headers for Publix Tab
output_sheet = wb_out.get_sheet_by_name("Publix")
output_sheet['A1'] = "Item"
output_sheet['B1'] = "Item UPC"
output_sheet['C1'] = "Style #"
output_sheet['D1'] = "Publix Unit Cost"
output_sheet['E1'] = "Publix Case Cost"
output_sheet['F1'] = "Case Count"
output_sheet['G1'] = "# of Single Units Returned"
output_sheet['H1'] = "# of Cases Returnd"
output_sheet['I1'] = "Return Total Unit Cost"
output_sheet['J1'] = "Return Total Case Cost"
output_sheet['K1'] = "Total Credit Amt (Single & Cases"
output_sheet['L1'] = "Supplier"
output_sheet['M1'] = "Retail Chain"
output_sheet['N1'] = "Return Date Reported"
output_sheet['O1'] = "Reason for Return"
output_sheet['P1'] = "Return Ticket #:"
output_sheet['Q1'] = "Store #"
output_sheet['R1'] = "Account #"

# add headers for Publix Tab
others_output_row = 1
ws_others = wb_out.create_sheet("Others", 0)
output_sheet = wb_out.get_sheet_by_name("Others")
output_sheet['A1'] = "Item"
output_sheet['B1'] = "Item UPC"
output_sheet['C1'] = "Kroger Unit Cost"
output_sheet['D1'] = "Kroger Case Cost"
output_sheet['E1'] = "Case Count"
output_sheet['F1'] = "# of Single Units Returned"
output_sheet['G1'] = "# of Cases REturnd"
output_sheet['H1'] = "Return Total Unit Coast"
output_sheet['I1'] = "Return Total Case Cost"
output_sheet['J1'] = "Total Credit Amt (Single & Cases"
output_sheet['K1'] = "Supplier"
output_sheet['L1'] = "Retail Chain"
output_sheet['M1'] = "Return Date Reported"
output_sheet['N1'] = "Reason for Return"
output_sheet['O1'] = "Return Ticket #:"
output_sheet['P1'] = "Store #"
output_sheet['Q1'] = "Account #"

logging.info("Output workbook created in memory")

upload_header_detail = []
upload_row_detail = []
upload_file_count = 0
upload_row_count = 0

# Open and read list of directories / programs that need to be processed
wb_mapping = openpyxl.load_workbook(os.path.join(output_directory, "Return_Ticket_Mapping.xlsx"))
mapping_ws = wb_mapping.get_active_sheet()
for row in range(2, mapping_ws.max_row+1):
    retailer = mapping_ws['A{}'.format(row)].value
    program = mapping_ws['B{}'.format(row)].value
    print("Retailer: {}".format(retailer) + "\n" + "Program: {}".format(program))
    logging.info("Retailer: {}     Program: {}".format(retailer, program))

    if program is None:
        directory = os.path.join("X:\\", retailer)
    else:
        directory = os.path.join("X:\\", retailer, program)

    # list off the files in each directory
    files = os.listdir(directory)
    for file in files:
        if os.path.isfile(os.path.join(directory, file)):
            upload_row_count = 0

            # process each file
            print("\n" + os.path.join(directory, file))
            logging.info("Current file: {}". format(os.path.join(directory, file)))
            with open(os.path.join(directory, file), 'rb') as f:
                memory_wb = io.BytesIO(f.read())
            wb = load_workbook(memory_wb, data_only=True, read_only=True)
            sheet = wb.get_active_sheet()

            # Publix
            if "publix" in file.lower():
                output_sheet = wb_out.get_sheet_by_name("Publix")
                detail_row = 16

                # check for empty row at the top of the file
                if sheet['B2'].value is not None:
                    detail_row = 17
                    # gather return details
                    supplier = sheet['C8'].value
                    chain = sheet['D2'].value
                    claim_month = sheet['C4'].value
                    claim_day = sheet['D4'].value
                    claim_year = sheet['E4'].value
                    reason = sheet['C7'].value
                    store = sheet['C9'].value
                    account = sheet['C11'].value
                else:
                    # gather return details
                    supplier = sheet['C7'].value
                    chain = sheet['D1'].value
                    claim_month = sheet['C3'].value
                    claim_day = sheet['D3'].value
                    claim_year = sheet['E3'].value
                    reason = sheet['C6'].value
                    store = sheet['C8'].value
                    account = sheet['C10'].value

                # if the template does not match the expected format log the errors
                if supplier == None or reason == None or \
                        store == None or \
                        claim_month == None or \
                        claim_day == None or \
                        claim_year== None:
                    errors.append(file)
                    wb.close()
                    time.sleep(1)
                    continue

                ret_ticket_num = supplier + reason + str(store) + str(claim_month) + str(claim_day) + str(claim_year)
                
                if reason.lower() == "dmg" or reason.lower() == "exp":
                    upload_file_count += 1
                    upload_header_detail.append([upload_file_count,
                                                 "dDocument_Items",
                                                 "{}{}{}".format(stmt_year, stmt_month, stmt_day),  # claim date
                                                 account,
                                                 ret_ticket_num,
                                                 "RTA Credit for {}".format(ret_ticket_num),  # claim description
                                                 "20{}{}{}".format(claim_year, claim_month, claim_day),  # claim date
                                                 0,
                                                 "N",
                                                 "RETURN"])

                for row in range(detail_row, sheet.max_row):
                    if sheet['B{}'.format(str(row))].value.lower() == 'totals':
                        break
                    else:
                        if sheet['L{}'.format(str(row))].value == 0:
                            continue # do not compile $0 rows

                        publix_output_row += 1
                        # Gather columns B - L and post to columns A - K for the output file
                        output_sheet['A{}'.format(str(publix_output_row))] = sheet['B{}'.format(str(row))].value
                        output_sheet['B{}'.format(str(publix_output_row))] = sheet['C{}'.format(str(row))].value
                        output_sheet['C{}'.format(str(publix_output_row))] = sheet['D{}'.format(str(row))].value
                        output_sheet['D{}'.format(str(publix_output_row))] = sheet['E{}'.format(str(row))].value
                        output_sheet['E{}'.format(str(publix_output_row))] = sheet['F{}'.format(str(row))].value
                        output_sheet['F{}'.format(str(publix_output_row))] = sheet['G{}'.format(str(row))].value
                        output_sheet['G{}'.format(str(publix_output_row))] = sheet['H{}'.format(str(row))].value
                        output_sheet['H{}'.format(str(publix_output_row))] = sheet['I{}'.format(str(row))].value
                        output_sheet['I{}'.format(str(publix_output_row))] = sheet['J{}'.format(str(row))].value
                        output_sheet['J{}'.format(str(publix_output_row))] = sheet['K{}'.format(str(row))].value
                        output_sheet['K{}'.format(str(publix_output_row))] = sheet['L{}'.format(str(row))].value
                        # Calculated fields
                        output_sheet['L{}'.format(str(publix_output_row))] = supplier
                        output_sheet['M{}'.format(str(publix_output_row))] = chain
                        output_sheet['N{}'.format(str(publix_output_row))] = \
                            str(claim_month) + "/" + str(claim_day) + "/" + str(claim_year)
                        output_sheet['O{}'.format(str(publix_output_row))] = reason
                        output_sheet['P{}'.format(str(publix_output_row))] = ret_ticket_num
                        output_sheet['Q{}'.format(str(publix_output_row))] = store
                        output_sheet['R{}'.format(str(publix_output_row))] = account
                        output_sheet['S{}'.format(str(publix_output_row))] = os.path.join(directory, file)

                        if reason.lower() == "dmg" or reason.lower() == "exp":
                            upload_row_count += 1
                            # lookup the item upc and return the item code
                            item_upc = sheet['C{}'.format(str(row))].value.strip()
                            item_code = 0
                            for i in product_lookup:
                                if item_upc == i[0]:
                                    item_code = i[1]
                                    logging.info("UPC: {} Mapped to: {}".format(item_upc, item_code))

                            upload_row_detail.append([upload_file_count,
                                                      upload_row_count,
                                                      item_code,
                                                      sheet['H{}'.format(str(row))].value,
                                                      "Y",
                                                      "TAXNO",
                                                      "Y",
                                                      sheet['E{}'.format(str(row))].value,
                                                      "MISAR",
                                                      "_SYS00000000254"])

            # All other accounts
            else:
                output_sheet = wb_out.get_sheet_by_name("Others")
                detail_row = 16

                if sheet['B2'].value is not None:
                    detail_row = 17
                    supplier = sheet['C8'].value
                    chain = sheet['D2'].value
                    claim_month = sheet['C4'].value
                    claim_day = sheet['D4'].value
                    claim_year = sheet['E4'].value
                    reason = sheet['C7'].value
                    store = sheet['C9'].value
                    account = sheet['C11'].value
                else:
                    supplier = sheet['C7'].value
                    chain = sheet['D1'].value
                    claim_month = sheet['C3'].value
                    claim_day = sheet['D3'].value
                    claim_year = sheet['E3'].value
                    reason = sheet['C6'].value
                    store = sheet['C8'].value
                    account = sheet['C10'].value

                if supplier == None or \
                        reason == None or \
                        store == None or \
                        claim_month == None or \
                        claim_day == None or \
                        claim_year == None:
                    errors.append(file)
                    wb.close()
                    time.sleep(1)
                    continue

                ret_ticket_num = str(supplier) + str(reason) + str(store) + str(claim_month) + str(claim_day) + \
                                 str(claim_year)

                if reason.lower() == "dmg" or reason.lower() == "exp":
                    upload_file_count += 1
                    upload_header_detail.append([upload_file_count,
                                                 "dDocument_Items",
                                                 "{}{}{}".format(stmt_year, stmt_month, stmt_day),
                                                 account,
                                                 ret_ticket_num,
                                                 "RTA Credit for {}".format(ret_ticket_num),  # claim description
                                                 "20{}{}{}".format(claim_year, claim_month, claim_day),  # claim date
                                                 0,
                                                 "N",
                                                 "RETURN"])

                # extract detail rows
                for row in range(detail_row, sheet.max_row):
                    if sheet['B{}'.format(str(row))].value == None or \
                            sheet['A{}'.format(str(row))].value == None:
                        continue
                    else:
                        if sheet['K{}'.format(str(row))].value == 0:
                            continue # do not compile $0 rows

                        others_output_row += 1
                        output_sheet['A{}'.format(str(others_output_row))] = sheet['B{}'.format(str(row))].value
                        output_sheet['B{}'.format(str(others_output_row))] = sheet['C{}'.format(str(row))].value
                        output_sheet['C{}'.format(str(others_output_row))] = sheet['D{}'.format(str(row))].value
                        output_sheet['D{}'.format(str(others_output_row))] = sheet['E{}'.format(str(row))].value
                        output_sheet['E{}'.format(str(others_output_row))] = sheet['F{}'.format(str(row))].value
                        output_sheet['F{}'.format(str(others_output_row))] = sheet['G{}'.format(str(row))].value
                        output_sheet['G{}'.format(str(others_output_row))] = sheet['H{}'.format(str(row))].value
                        output_sheet['H{}'.format(str(others_output_row))] = sheet['I{}'.format(str(row))].value
                        output_sheet['I{}'.format(str(others_output_row))] = sheet['J{}'.format(str(row))].value
                        output_sheet['J{}'.format(str(others_output_row))] = sheet['K{}'.format(str(row))].value
                        output_sheet['K{}'.format(str(others_output_row))] = supplier
                        output_sheet['L{}'.format(str(others_output_row))] = chain
                        output_sheet['M{}'.format(str(others_output_row))] = \
                            str(claim_month) + "/" + str(claim_day) + "/" + str(claim_year)
                        output_sheet['N{}'.format(str(others_output_row))] = reason
                        output_sheet['O{}'.format(str(others_output_row))] = ret_ticket_num
                        output_sheet['P{}'.format(str(others_output_row))] = store
                        output_sheet['Q{}'.format(str(others_output_row))] = account
                        output_sheet['R{}'.format(str(others_output_row))] = os.path.join(directory, file)

                        if reason.lower() == "dmg" or reason.lower() == "exp":
                            upload_row_count += 1
                            # lookup the item upc and return the item code
                            item_upc = sheet['C{}'.format(str(row))].value.strip()
                            item_code = 0
                            for i in product_lookup:
                                if item_upc == i[0]:
                                    item_code = i[1]
                                    logging.info("UPC: {} Mapped to: {}".format(item_upc, item_code))

                            upload_row_detail.append([upload_file_count,
                                                      upload_row_count,
                                                      item_code,
                                                      sheet['G{}'.format(str(row))].value,
                                                      "Y",
                                                      "TAXNO",
                                                      "Y",
                                                      sheet['D{}'.format(str(row))].value,
                                                      "MISAR",
                                                      "_SYS00000000254"])
            wb.close()
            time.sleep(1)  # allow a second for the file to close
            # test path
            logging.info("Source: {}".format(os.path.join(directory, file)))
            logging.info("Destination: {}".format(os.path.join(output_directory, "Entered On Tracker")))

            # Test
            # shutil.copy2(os.path.join(directory, file), os.path.join(output_directory, "Entered On Tracker"))

            # Production
            try:
                shutil.copy2(os.path.join(directory, file), os.path.join(directory, "Entered On Tracker"))
                print()
            except shutil.SameFileError as e:
                print(e)
                logging.warning("Duplicate file")
                logging.warning(e)

            # if os.path.isfile(os.path.join(output_directory, "Entered On Tracker", file)):
            if os.path.isfile(os.path.join(directory, "Entered On Tracker", file)):
                print("File copied successfully")
                logging.info("File successfully saved to destination folder")
                files_processed.append(os.path.join(directory, file))
            else:
                logging.warning("Destination file was not found")

# open dummy workbook to release the last file processed
# wb = load_workbook(os.path.join("L:\\SAP_Return_Tickets", "Empty_Workbook.xlsx"), data_only=True, read_only=True)
# sheet = wb.get_active_sheet()
# wb.close()


wb_out.save(os.path.join(output_file_directory, "Return_Tickets_{}.{}.{}.xlsx".format(month, day, year)))
logging.info("Output file saved")
wb_mapping.close()
logging.info("Process completed")

output_file = open(os.path.join(output_directory, "files_processed.csv"), 'w', newline='')
csv_writer = csv.writer(output_file)

for file in files_processed:
    logging.info(file)
    csv_writer.writerow([file])
    try:
        os.remove(file)
        time.sleep(10)
    except PermissionError as e:
        logging.warning("Could not delete {}".format(file))
        logging.warning(e)
    if os.path.isfile(file):
        logging.info("File not deleted")
    else:
        logging.info("{} successfully deleted".format(file))
    print(file)
output_file.close()

upload_header_file = open(os.path.join(output_file_directory,
                                       "upload_header_{}.{}.{}.csv".format(month, day, year)),
                                        'w',
                                        newline='')
csv_writer = csv.writer(upload_header_file)
csv_writer.writerow(["DocNum",
                     "DocType",
                     "DocDate",
                     "CardCode",
                     "NumAtCard",
                     "Comments",
                     "TaxDate",
                     "U_A1WMS_ADJ",
                     "U_EDISend",
                     "U_ReturnReasonCode"])
csv_writer.writerow(["DocumentNumber",
                     "DocumentType",
                     "Document Date",
                     "Card Code",
                     "NumAtCard",
                     "Comments",
                     "TaxDate",
                     "A1WMS_ADJ",
                     "EDISend",
                     "Return Reason Code"])
for row in upload_header_detail:
    csv_writer.writerow(row)
upload_header_file.close()

upload_row_file = open(os.path.join(output_file_directory, "upload_row_detail_{}.{}.{}.csv".format(month, day, year)), 'w',
                          newline='')
csv_writer = csv.writer(upload_row_file)
csv_writer.writerow(["ParentKey",
                     "LineNum",
                     "ItemCode",
                     "Quantity",
                     "UseBaseUnits",
                     "TaxCode",
                     "WithoutInventoryMovement",
                     "UnitPrice",
                     "WarehouseCode",
                     "AccountCode"])
csv_writer.writerow(["DocNum",
                     "LineNum",
                     "ItemCode",
                     "Quantity",
                     "UseBaseUnits",
                     "TaxCode",
                     "NoInvtryMv",
                     "PriceBegDi",
                     "WhsCode",
                     "AcctCode"])
for row in upload_row_detail:
    csv_writer.writerow(row)
upload_row_file.close()

output_files = [
    os.path.join(output_file_directory, "Return_Tickets_{}.{}.{}.xlsx".format(month, day, year)),
    os.path.join(output_file_directory, "upload_header_{}.{}.{}.csv".format(month, day, year)),
    os.path.join(output_file_directory, "upload_row_detail_{}.{}.{}.csv".format(month, day, year))
]
send_email.send_message(output_files, errors)